import streamlit as st
import pandas as pd
import sys
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append('.')
from utils.database import load_all_logs, get_premium_status

st.set_page_config(page_title="Export & History | The Lymphie Sanctuary", page_icon="📊", layout="wide")

from utils.nav import mobile_nav
mobile_nav()

from utils.styles import apply_styles
apply_styles()

st.title("📊 Export & History")
st.caption("View your past entries and download your data.")

premium = get_premium_status()
logs = load_all_logs()

if not logs:
    st.info("📝 No entries yet. Your daily logs will appear here once you start tracking.")
    if st.button("📝 Go to Daily Log", use_container_width=True):
        st.switch_page("pages/2_Daily_Log.py")
    st.stop()

df = pd.DataFrame(logs)

st.subheader("📋 Your Entries")
st.dataframe(df, use_container_width=True, hide_index=True)
st.caption(f"📊 **{len(logs)}** total entries logged")

st.markdown("---")
st.subheader("💾 Backup Reminder")
st.markdown("""
<div style="background: var(--tint); padding: 1.2rem; border-radius: 12px; border-left: 4px solid var(--teal);">
    <strong>📥 Your data is stored locally in a database file.</strong><br>
    Download an Excel backup regularly and keep it somewhere safe.
</div>
""", unsafe_allow_html=True)

st.markdown("---")
st.subheader("📥 Export Your Data")

if premium:
    df_export = df.copy()
    cols_to_drop = ['id', 'created_at']
    df_export = df_export.drop(columns=[col for col in cols_to_drop if col in df_export.columns])

    column_rename = {
        'date': 'Date',
        'time': 'Time',
        'heaviness': 'Heaviness (0-10)',
        'pain': 'Pain (0-10)',
        'limb_appearance': 'Limb Appearance',
        'measurement_taken': 'Measurement Taken',
        'affected_areas': 'Affected Areas',
        'compression_type': 'Compression Worn',
        'compression_hours': 'Compression Hours',
        'self_care': 'Home Self-Care',
        'professional_treatment': 'Professional Treatment',
        'movement_exercise': 'Movement & Exercise',
        'dietary_triggers': 'Dietary Triggers',
        'environmental_triggers': 'Environmental Triggers',
        'health_triggers': 'Health Triggers',
        'stress': 'Stress (0-10)',
        'sleep_quality': 'Sleep Quality',
        'energy': 'Energy (0-10)',
        'mobility': 'Mobility (0-10)',
        'self_compassion': 'Self-Compassion (0-10)',
        'biggest_challenge': 'Biggest Challenge',
        'small_win': 'Small Win',
        'temperature': 'Temperature (°C)',
        'humidity': 'Humidity (%)',
        'tags': 'Tags'
    }
    df_export = df_export.rename(
        columns={k: v for k, v in column_rename.items() if k in df_export.columns}
    )
    df_export = df_export.fillna('')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Lymphie Sanctuary Logs')
        workbook = writer.book
        worksheet = writer.sheets['Lymphie Sanctuary Logs']

        # ── Styles ──────────────────────────────────────────────
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_alignment = Alignment(vertical='top', wrap_text=True)
        cell_font = Font(name='Calibri', size=10, color='000000')

        black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        column_widths = {
            'Date': 12, 'Time': 8, 'Heaviness (0-10)': 14, 'Pain (0-10)': 12,
            'Limb Appearance': 28, 'Measurement Taken': 22, 'Affected Areas': 30,
            'Compression Worn': 28, 'Compression Hours': 16, 'Home Self-Care': 30,
            'Professional Treatment': 30, 'Movement & Exercise': 28,
            'Dietary Triggers': 28, 'Environmental Triggers': 30, 'Health Triggers': 28,
            'Stress (0-10)': 12, 'Sleep Quality': 18, 'Energy (0-10)': 12,
            'Mobility (0-10)': 14, 'Self-Compassion (0-10)': 18,
            'Biggest Challenge': 45, 'Small Win': 45, 'Temperature (°C)': 16,
            'Humidity (%)': 14, 'Tags': 25
        }

        for col_idx in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = black_border

            col_name = df_export.columns[col_idx - 1]
            if col_name in column_widths:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = column_widths[col_name]
            else:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 20

        for row_idx in range(2, len(df_export) + 2):
            for col_idx in range(1, len(df_export.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = cell_alignment
                cell.border = black_border
                cell.font = cell_font

        worksheet.row_dimensions[1].height = 30
        for row_idx in range(2, len(df_export) + 2):
            worksheet.row_dimensions[row_idx].height = 22

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df_export.columns))}{len(df_export) + 1}"

    output.seek(0)
    excel_data = output.read()

    st.download_button(
        label="📥 Download Beautiful Excel Backup (.xlsx)",
        data=excel_data,
        file_name=f"Lymphie_Sanctuary_Export_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.success("✅ Lifetime Access active — export anytime.")
    st.caption("✨ Formatted, professional, and ready to take to your next appointment. Show your therapist what your body has actually been doing.")

else:
    st.warning("✨ Full Export requires Lifetime Access")
    st.markdown("""
    <div style="background: var(--tint); padding: 1.5rem; border-radius: 16px; margin: 1rem 0; text-align: center;">
        <h3 style="color: #1A3B2E; margin-bottom: 0.5rem; font-family: 'Nunito', sans-serif;">NZ$19.99 — One‑time payment</h3>
        <p>Lifetime access to formatted Excel exports, trend visualizations, and all future updates.</p>
    </div>
    """, unsafe_allow_html=True)
    if st.button("🌿 Go to Settings to Unlock", use_container_width=True):
        st.switch_page("pages/1_Settings.py")

# ─── BRAND FOOTER ────────────────────────────────────────────────
st.divider()
st.markdown("""
<div class="brand-footer">
    Does your workplace support staff with lymphoedema?
    <a href="https://www.lymphatwork.com" target="_blank">Lymphoedema at Work →</a>
</div>
""", unsafe_allow_html=True)